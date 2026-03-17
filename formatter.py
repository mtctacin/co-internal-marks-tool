import pandas as pd
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib.pagesizes import A4


def normalize_marks(file_path):

    df = pd.read_excel(file_path)

    raw_max = df.iloc[0]
    norm_max = df.iloc[1]

    students = df.iloc[2:].reset_index(drop=True)

    co_map = {}
    eval_methods = {}

    for col in df.columns:

        if "_CO" in col:

            comp, co = col.split("_")

            co_map.setdefault(co, []).append(col)
            eval_methods.setdefault(co, []).append(comp)

    co_columns = sorted(co_map.keys())

    normalized = pd.DataFrame()
    normalized["RegNo"] = students["RegNo"]
    normalized["Name"] = students["Name"]

    max_marks = []

    for co in co_columns:

        total = 0

        for col in co_map[co]:
            total += (students[col] / raw_max[col]) * norm_max[col]

        normalized[co] = total.round(2)

        max_marks.append(sum(norm_max[c] for c in co_map[co]))

    eval_list = [", ".join(eval_methods[co]) for co in co_columns]

    return normalized, co_columns, eval_list, max_marks


def generate_part_tables(df, co_columns):

    part2 = df.copy()

    part2["Total Marks Obtained in CCA"] = part2[co_columns].sum(axis=1).round(2)

    part2.insert(0, "SL. No.", range(1, len(part2)+1))

    part2.rename(columns={
        "RegNo": "Register Number",
        "Name": "Name of the Student"
    }, inplace=True)

    for co in co_columns:
        part2.rename(columns={co: f"Marks Obtained in {co}"}, inplace=True)

    part1 = df.copy()

    part1.insert(0, "SL. No.", range(1, len(part1)+1))

    part1.rename(columns={
        "RegNo": "Register Number",
        "Name": "Name of the Student"
    }, inplace=True)

    part1["Total Marks Obtained in CCA"] = part1[co_columns].sum(axis=1).round(2)

    for co in co_columns:
        part1.rename(columns={co: f"Marks Obtained in {co}"}, inplace=True)

    return part1, part2


def export_excel_template(
    part1,
    part2,
    college,
    course_code,
    course_name,
    semester,
    year,
    template_path,
    output_file
):

    wb = load_workbook(template_path)

    ws1 = wb["CCA Form A Part I"]
    ws2 = wb["CCA Form A Part II"]

    # ---------- HEADER ----------
    ws1["C2"] = college
    ws1["L2"] = year
    ws1["C3"] = course_code
    ws1["K3"] = course_name
    ws1["C4"] = semester

    ws2["C2"] = college
    ws2["L2"] = year
    ws2["C3"] = course_code
    ws2["K3"] = course_name
    ws2["C4"] = semester

    # ---------- PART I DATA ----------
    start_row = 10

    for i, row in part1.iterrows():

        r = start_row + i

        ws1.cell(r,1,row["SL. No."])
        ws1.cell(r,2,row["Register Number"])
        ws1.cell(r,3,row["Name of the Student"])

        col_index = 4

        for col in part1.columns[3:]:

            ws1.cell(r,col_index,row[col])
            col_index += 1

    # ---------- PART II DATA ----------

    start_row = 10

    for i, row in part2.iterrows():

        r = start_row + i

        ws2.cell(r,1,row["SL. No."])
        ws2.cell(r,2,row["Register Number"])
        ws2.cell(r,3,row["Name of the Student"])

        col_index = 4

        for col in part2.columns[3:]:

            ws2.cell(r,col_index,row[col])
            col_index += 1

    wb.save(output_file)


def export_pdf(part2_df, output_file):

    data = [list(part2_df.columns)]

    for _, row in part2_df.iterrows():
        data.append(list(row))

    pdf = SimpleDocTemplate(output_file, pagesize=A4)

    table = Table(data)

    pdf.build([table])
